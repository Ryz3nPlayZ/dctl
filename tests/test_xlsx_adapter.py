from __future__ import annotations

from pathlib import Path
import tempfile
import unittest

from openpyxl import Workbook

from dctl.adapters import xlsx_files


class XlsxAdapterTests(unittest.TestCase):
    def test_read_and_write_cell(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "sample.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Sheet1"
            sheet["A1"] = "name"
            sheet["B1"] = "count"
            workbook.save(path)

            write_result = xlsx_files.write_cell(str(path), "Sheet1", "B2", "42", json_value=True)
            result = xlsx_files.read(str(path), "Sheet1", "A1:B2")

            self.assertTrue(Path(write_result["backup_path"]).exists())
            self.assertEqual(result["values"], [["name", "count"], [None, 42]])

    def test_write_range(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "sample.xlsx"
            workbook = Workbook()
            workbook.active.title = "Sheet1"
            workbook.save(path)

            write_result = xlsx_files.write_range(str(path), "Sheet1", "A1:B2", '[["a", 1], ["b", 2]]')
            result = xlsx_files.read(str(path), "Sheet1", "A1:B2")

            self.assertTrue(Path(write_result["backup_path"]).exists())
            self.assertEqual(result["values"], [["a", 1], ["b", 2]])

    def test_worksheet_map_and_locate_fill_by_labels(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "worksheet.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Sheet1"
            sheet["A1"] = ""
            sheet["B1"] = "Definition"
            sheet["C1"] = "Example"
            sheet["A2"] = "Mitochondria"
            sheet["A3"] = "Chlorophyll"
            workbook.save(path)

            worksheet = xlsx_files.worksheet_map(str(path), sheet_name="Sheet1")
            self.assertEqual(worksheet["items"][0]["inferred_table"]["header_row"], ["Definition", "Example"])

            located = xlsx_files.locate_cell(str(path), "Sheet1", "Mitochondria", "Definition")
            self.assertEqual(located["cell"], "B2")

            filled = xlsx_files.fill_cell(str(path), "Sheet1", "Chlorophyll", "Example", "Green pigment")
            self.assertTrue(Path(filled["backup_path"]).exists())
            self.assertEqual(xlsx_files.read(str(path), "Sheet1", "A1:C3")["values"][2][2], "Green pigment")

    def test_fill_table_batch(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            path = Path(tmpdir) / "worksheet.xlsx"
            workbook = Workbook()
            sheet = workbook.active
            sheet.title = "Sheet1"
            sheet["A1"] = ""
            sheet["B1"] = "Atomic Number"
            sheet["C1"] = "Symbol"
            sheet["A2"] = "Hydrogen"
            sheet["A3"] = "Oxygen"
            workbook.save(path)

            result = xlsx_files.fill_table(
                str(path),
                "Sheet1",
                '[{"row_label":"Hydrogen","column_label":"Atomic Number","value":1},{"row_label":"Oxygen","column_label":"Symbol","value":"O"}]',
            )
            self.assertEqual(len(result["items"]), 2)
            values = xlsx_files.read(str(path), "Sheet1", "A1:C3")["values"]
            self.assertEqual(values[1][1], 1)
            self.assertEqual(values[2][2], "O")


if __name__ == "__main__":
    unittest.main()
