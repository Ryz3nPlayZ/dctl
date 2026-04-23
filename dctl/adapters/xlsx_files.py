from __future__ import annotations

import difflib
import json
from pathlib import Path
import shutil
from typing import Any

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell
from openpyxl.utils.cell import get_column_letter, range_boundaries

from dctl.errors import DctlError


BACKUP_DIRNAME = ".dctl-backups"


def _path(path: str) -> Path:
    resolved = Path(path).expanduser().resolve()
    if not resolved.exists():
        raise DctlError("ELEMENT_NOT_FOUND", f"Workbook not found: {resolved}")
    return resolved


def _normalize_text(value: Any) -> str:
    return " ".join(str(value or "").split()).strip().casefold()


def _parse_json_input(value: str) -> Any:
    candidate = Path(value).expanduser()
    raw = candidate.read_text(encoding="utf-8") if candidate.exists() else value
    try:
        return json.loads(raw)
    except json.JSONDecodeError as exc:
        raise DctlError("INVALID_SELECTOR", f"Expected JSON or a JSON file path: {value}") from exc


def _backup_path(path: Path) -> Path:
    backup_dir = path.parent / BACKUP_DIRNAME
    backup_dir.mkdir(exist_ok=True)
    counter = 1
    while True:
        candidate = backup_dir / f"{path.stem}.backup-{counter}{path.suffix}"
        if not candidate.exists():
            return candidate
        counter += 1


def backup(path: str) -> dict[str, Any]:
    workbook_path = _path(path)
    destination = _backup_path(workbook_path)
    shutil.copy2(workbook_path, destination)
    return {"path": str(workbook_path), "backup_path": str(destination)}


def _sheet(workbook: Any, name: str) -> Any:
    if name not in workbook.sheetnames:
        raise DctlError(
            "ELEMENT_NOT_FOUND",
            f"Worksheet '{name}' was not found.",
            details={"sheets": workbook.sheetnames},
        )
    return workbook[name]


def inspect(path: str) -> dict[str, Any]:
    workbook_path = _path(path)
    workbook = load_workbook(workbook_path)
    return {
        "path": str(workbook_path),
        "sheet_count": len(workbook.sheetnames),
        "active_sheet": workbook.active.title,
        "sheets": workbook.sheetnames,
    }


def sheets(path: str) -> dict[str, Any]:
    workbook_path = _path(path)
    workbook = load_workbook(workbook_path)
    items = []
    for name in workbook.sheetnames:
        worksheet = workbook[name]
        items.append(
            {
                "name": name,
                "max_row": worksheet.max_row,
                "max_column": worksheet.max_column,
                "tables": list(worksheet.tables.keys()),
            }
        )
    return {"path": str(workbook_path), "items": items}


def read(path: str, sheet_name: str, cell_range: str) -> dict[str, Any]:
    workbook_path = _path(path)
    workbook = load_workbook(workbook_path, data_only=False)
    worksheet = _sheet(workbook, sheet_name)
    values = [[cell.value for cell in row] for row in worksheet[cell_range]]
    return {"path": str(workbook_path), "sheet": sheet_name, "range": cell_range, "values": values}


def _coerce_value(raw: str, json_value: bool) -> Any:
    if not json_value:
        return raw
    try:
        return json.loads(raw)
    except json.JSONDecodeError as exc:
        raise DctlError("INVALID_SELECTOR", f"Value is not valid JSON: {raw}") from exc


def _write_value(cell: Cell, value: Any) -> None:
    if isinstance(value, str) and value.startswith("="):
        cell.value = value
        return
    cell.value = value


def write_cell(path: str, sheet_name: str, cell_ref: str, raw_value: str, json_value: bool = False) -> dict[str, Any]:
    workbook_path = _path(path)
    workbook = load_workbook(workbook_path)
    worksheet = _sheet(workbook, sheet_name)
    value = _coerce_value(raw_value, json_value)
    backup_result = backup(str(workbook_path))
    _write_value(worksheet[cell_ref], value)
    workbook.save(workbook_path)
    return {"path": str(workbook_path), "sheet": sheet_name, "cell": cell_ref, "value": value, "backup_path": backup_result["backup_path"]}


def write_range(path: str, sheet_name: str, cell_range: str, json_rows: str) -> dict[str, Any]:
    workbook_path = _path(path)
    workbook = load_workbook(workbook_path)
    worksheet = _sheet(workbook, sheet_name)
    try:
        rows = json.loads(json_rows)
    except json.JSONDecodeError as exc:
        raise DctlError("INVALID_SELECTOR", "Range data must be valid JSON.") from exc
    if not isinstance(rows, list) or not all(isinstance(row, list) for row in rows):
        raise DctlError("INVALID_SELECTOR", "Range data must be a JSON array of arrays.")
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    expected_rows = max_row - min_row + 1
    expected_cols = max_col - min_col + 1
    if len(rows) != expected_rows or any(len(row) != expected_cols for row in rows):
        raise DctlError(
            "INVALID_SELECTOR",
            f"Range {cell_range} expects {expected_rows}x{expected_cols} values.",
        )
    backup_result = backup(str(workbook_path))
    for row_offset, row in enumerate(rows):
        for col_offset, value in enumerate(row):
            cell = worksheet.cell(row=min_row + row_offset, column=min_col + col_offset)
            _write_value(cell, value)
    workbook.save(workbook_path)
    return {"path": str(workbook_path), "sheet": sheet_name, "range": cell_range, "rows": rows, "backup_path": backup_result["backup_path"]}


def _sheet_rows(worksheet: Any) -> list[list[Any]]:
    return [
        [worksheet.cell(row=row_index, column=column_index).value for column_index in range(1, worksheet.max_column + 1)]
        for row_index in range(1, worksheet.max_row + 1)
    ]


def _first_non_empty_row(rows: list[list[Any]]) -> int | None:
    for index, row in enumerate(rows, start=1):
        if any(value not in (None, "") for value in row):
            return index
    return None


def _table_descriptor_from_ref(worksheet: Any, table_name: str, ref: str) -> dict[str, Any]:
    min_col, min_row, max_col, max_row = range_boundaries(ref)
    header_row = [
        worksheet.cell(row=min_row, column=column).value
        for column in range(min_col + 1, max_col + 1)
    ]
    row_labels = [
        worksheet.cell(row=row, column=min_col).value
        for row in range(min_row + 1, max_row + 1)
    ]
    return {
        "name": table_name,
        "ref": ref,
        "header_row": header_row,
        "row_labels": row_labels,
        "sheet": worksheet.title,
    }


def _inferred_table_descriptor(worksheet: Any) -> dict[str, Any]:
    rows = _sheet_rows(worksheet)
    header_row_index = _first_non_empty_row(rows)
    if header_row_index is None:
        return {
            "name": None,
            "ref": None,
            "header_row": [],
            "row_labels": [],
            "sheet": worksheet.title,
        }
    header_values = rows[header_row_index - 1]
    first_data_column = 1
    first_row_label_column = 1
    if header_values and header_values[0] in (None, ""):
        first_data_column = 2
    header_row = [value for value in header_values[first_data_column - 1:] if value not in (None, "")]
    row_labels = [
        row[first_row_label_column - 1]
        for row in rows[header_row_index:]
        if row and row[first_row_label_column - 1] not in (None, "")
    ]
    return {
        "name": "__inferred__",
        "ref": f"A{header_row_index}:{get_column_letter(max(worksheet.max_column, 1))}{max(worksheet.max_row, header_row_index)}",
        "header_row": header_row,
        "row_labels": row_labels,
        "sheet": worksheet.title,
    }


def worksheet_map(path: str, sheet_name: str | None = None) -> dict[str, Any]:
    workbook_path = _path(path)
    workbook = load_workbook(workbook_path)
    sheets_to_emit = [sheet_name] if sheet_name else workbook.sheetnames
    items = []
    for name in sheets_to_emit:
        worksheet = _sheet(workbook, name)
        tables = [_table_descriptor_from_ref(worksheet, table_name, table.ref) for table_name, table in worksheet.tables.items()]
        inferred = _inferred_table_descriptor(worksheet)
        items.append(
            {
                "name": worksheet.title,
                "max_row": worksheet.max_row,
                "max_column": worksheet.max_column,
                "tables": tables,
                "inferred_table": inferred,
            }
        )
    return {"path": str(workbook_path), "items": items}


def _match_text_index(values: list[Any], selector: str, label: str) -> int:
    normalized = _normalize_text(selector)
    exact_matches = [index for index, value in enumerate(values) if _normalize_text(value) == normalized]
    if len(exact_matches) == 1:
        return exact_matches[0]
    partial_matches = [index for index, value in enumerate(values) if normalized and normalized in _normalize_text(value)]
    matches = exact_matches or partial_matches
    if not matches:
        raise DctlError("ELEMENT_NOT_FOUND", f"No {label} matching '{selector}' was found.")
    if len(matches) > 1:
        raise DctlError("MULTIPLE_MATCHES", f"{label.title()} selector '{selector}' matched multiple values.", details={label: matches})
    return matches[0]


def _resolve_table_descriptor(worksheet: Any, table_name: str | None = None) -> dict[str, Any]:
    if table_name:
        for name, table in worksheet.tables.items():
            if _normalize_text(name) == _normalize_text(table_name) or _normalize_text(table_name) in _normalize_text(name):
                return _table_descriptor_from_ref(worksheet, name, table.ref)
        raise DctlError(
            "ELEMENT_NOT_FOUND",
            f"No worksheet table matching '{table_name}' was found.",
            suggestion="Run `dctl xlsx worksheet-map` to inspect available tables.",
        )
    return _inferred_table_descriptor(worksheet)


def locate_cell(path: str, sheet_name: str, row_label: str, column_label: str, table_name: str | None = None) -> dict[str, Any]:
    workbook_path = _path(path)
    workbook = load_workbook(workbook_path)
    worksheet = _sheet(workbook, sheet_name)
    descriptor = _resolve_table_descriptor(worksheet, table_name=table_name)
    if not descriptor["ref"]:
        raise DctlError("ACTION_NOT_SUPPORTED", "Unable to infer a tabular region for this worksheet.")
    min_col, min_row, _max_col, _max_row = range_boundaries(descriptor["ref"])
    row_offset = _match_text_index(descriptor["row_labels"], row_label, "row_label") + 1
    column_offset = _match_text_index(descriptor["header_row"], column_label, "column_label") + 1
    row_index = min_row + row_offset
    column_index = min_col + column_offset
    cell = worksheet.cell(row=row_index, column=column_index)
    return {
        "path": str(workbook_path),
        "sheet": sheet_name,
        "table": descriptor["name"],
        "cell": cell.coordinate,
        "row_index": row_index,
        "column_index": column_index,
        "row_label": row_label,
        "column_label": column_label,
        "value": cell.value,
    }


def fill_cell(path: str, sheet_name: str, row_label: str, column_label: str, raw_value: str, *, table_name: str | None = None, json_value: bool = False) -> dict[str, Any]:
    workbook_path = _path(path)
    workbook = load_workbook(workbook_path)
    worksheet = _sheet(workbook, sheet_name)
    descriptor = _resolve_table_descriptor(worksheet, table_name=table_name)
    if not descriptor["ref"]:
        raise DctlError("ACTION_NOT_SUPPORTED", "Unable to infer a tabular region for this worksheet.")
    min_col, min_row, _max_col, _max_row = range_boundaries(descriptor["ref"])
    row_offset = _match_text_index(descriptor["row_labels"], row_label, "row_label") + 1
    column_offset = _match_text_index(descriptor["header_row"], column_label, "column_label") + 1
    row_index = min_row + row_offset
    column_index = min_col + column_offset
    value = _coerce_value(raw_value, json_value)
    backup_result = backup(str(workbook_path))
    cell = worksheet.cell(row=row_index, column=column_index)
    _write_value(cell, value)
    workbook.save(workbook_path)
    return {
        "path": str(workbook_path),
        "sheet": sheet_name,
        "table": descriptor["name"],
        "cell": cell.coordinate,
        "row_label": row_label,
        "column_label": column_label,
        "value": value,
        "backup_path": backup_result["backup_path"],
    }


def fill_table(path: str, sheet_name: str, entries_json: str, table_name: str | None = None) -> dict[str, Any]:
    payload = _parse_json_input(entries_json)
    if not isinstance(payload, list):
        raise DctlError("INVALID_SELECTOR", "Table fill payload must be a JSON array.")
    workbook_path = _path(path)
    workbook = load_workbook(workbook_path)
    worksheet = _sheet(workbook, sheet_name)
    descriptor = _resolve_table_descriptor(worksheet, table_name=table_name)
    if not descriptor["ref"]:
        raise DctlError("ACTION_NOT_SUPPORTED", "Unable to infer a tabular region for this worksheet.")
    min_col, min_row, _max_col, _max_row = range_boundaries(descriptor["ref"])
    backup_result = backup(str(workbook_path))
    results = []
    for item in payload:
        if not isinstance(item, dict) or not {"row_label", "column_label", "value"} <= set(item):
            raise DctlError("INVALID_SELECTOR", "Each table entry must include `row_label`, `column_label`, and `value`.")
        row_offset = _match_text_index(descriptor["row_labels"], str(item["row_label"]), "row_label") + 1
        column_offset = _match_text_index(descriptor["header_row"], str(item["column_label"]), "column_label") + 1
        cell = worksheet.cell(row=min_row + row_offset, column=min_col + column_offset)
        _write_value(cell, item["value"])
        results.append(
            {
                "row_label": item["row_label"],
                "column_label": item["column_label"],
                "cell": cell.coordinate,
                "value": item["value"],
            }
        )
    workbook.save(workbook_path)
    return {
        "path": str(workbook_path),
        "sheet": sheet_name,
        "table": descriptor["name"],
        "items": results,
        "backup_path": backup_result["backup_path"],
    }


def diff(path: str, against: str) -> dict[str, Any]:
    current_path = _path(path)
    against_path = _path(against)
    current_workbook = load_workbook(current_path, data_only=False)
    against_workbook = load_workbook(against_path, data_only=False)
    current_lines = []
    against_lines = []
    for sheet_name in current_workbook.sheetnames:
        worksheet = current_workbook[sheet_name]
        current_lines.append(f"[{sheet_name}]")
        for row in _sheet_rows(worksheet):
            current_lines.append(json.dumps(row, ensure_ascii=True))
    for sheet_name in against_workbook.sheetnames:
        worksheet = against_workbook[sheet_name]
        against_lines.append(f"[{sheet_name}]")
        for row in _sheet_rows(worksheet):
            against_lines.append(json.dumps(row, ensure_ascii=True))
    diff_lines = list(
        difflib.unified_diff(
            against_lines,
            current_lines,
            fromfile=str(against_path),
            tofile=str(current_path),
            lineterm="",
        )
    )
    return {"path": str(current_path), "against": str(against_path), "lines": diff_lines}
